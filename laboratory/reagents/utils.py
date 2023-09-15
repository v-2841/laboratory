import os

import openpyxl
from django.conf import settings
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, Side
from tempfile import NamedTemporaryFile

from reagents.models import Reagent


START_CELL_ROW = 4


def cell_style(cell):
    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )
    cell.border = Border(
        left=Side('thin'),
        right=Side('thin'),
        top=Side('thin'),
        bottom=Side('thin'),
    )
    cell.font = Font(name='Times New Roman', size=14)
    return cell


def cell_style_left(cell):
    cell.alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
    )
    cell.border = Border(
        left=Side('thin'),
        right=Side('thin'),
        top=Side('thin'),
        bottom=Side('thin'),
    )
    cell.font = Font(name='Times New Roman', size=14)
    return cell


def reagents_table_xlsx():
    workbook = openpyxl.load_workbook(
        os.path.join(settings.BASE_DIR, 'data', 'reagents_template.xlsx'))
    sheet = workbook.active

    reagents = Reagent.objects.all()
    cell_row = START_CELL_ROW
    for reagent in reagents:
        cell = sheet.cell(row=cell_row, column=1)
        cell = cell_style(cell)
        cell.value = reagent.index if reagent.index else '-'
        cell = sheet.cell(row=cell_row, column=2)
        cell = cell_style_left(cell)
        cell.value = reagent.name
        cell = sheet.cell(row=cell_row, column=3)
        cell = cell_style(cell)
        cell.value = reagent.grade if reagent.grade else '-'
        cell = sheet.cell(row=cell_row, column=4)
        cell = cell_style(cell)
        cell.value = reagent.amount if reagent.amount else '-'
        cell = sheet.cell(row=cell_row, column=5)
        cell = cell_style(cell)
        cell.value = reagent.manufacture_date.strftime(
            "%d.%m.%Y") if reagent.manufacture_date else '-'
        cell = sheet.cell(row=cell_row, column=6)
        cell = cell_style(cell)
        cell.value = reagent.expiration_date.strftime(
            "%d.%m.%Y") if reagent.expiration_date else '-'
        cell_row += 1

    with NamedTemporaryFile() as tmp:
        tmp.close()
        workbook.save(tmp.name)
        with open(tmp.name, 'rb') as f:
            f.seek(0)
            xlsx_data = f.read()
    response = HttpResponse(xlsx_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reagents_table.xlsx"'
    return response
