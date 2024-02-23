import locale
import os
from datetime import datetime
from time import time

import openpyxl
from django.conf import settings
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, Side
from tempfile import NamedTemporaryFile

from results.models import Result


START_CELL_ROW = 4


def cell_style(cell):
    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )
    cell.border = Border(
        left=Side('thin'),
        right=Side('thin'),
        top=Side('thin'),
        bottom=Side('thin'),
    )
    cell.font = Font(name='Times New Roman', size=12)
    return cell


def results_table_xlsx(marked_results=None):
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    workbook = openpyxl.load_workbook(
        os.path.join(settings.BASE_DIR, 'data', 'results_template.xlsx'))
    sheet = workbook.active

    if marked_results:
        results = Result.objects.filter(
            id__in=marked_results).prefetch_related('researcher')
    else:
        results = Result.objects.all().prefetch_related('researcher')
    cell_row = START_CELL_ROW
    for result in results:
        cell = sheet.cell(row=cell_row, column=1)
        cell = cell_style(cell)
        cell.value = result.pub_date.strftime("%d %B %Y г. %H:%M")
        cell = sheet.cell(row=cell_row, column=2)
        cell = cell_style(cell)
        cell.value = result.sample_name
        cell = sheet.cell(row=cell_row, column=3)
        cell = cell_style(cell)
        cell.value = result.analysis_name
        cell = sheet.cell(row=cell_row, column=4)
        cell = cell_style(cell)
        cell.value = result.standard
        cell = sheet.cell(row=cell_row, column=5)
        cell = cell_style(cell)
        cell.value = result.measurement_unit
        cell = sheet.cell(row=cell_row, column=6)
        cell = cell_style(cell)
        cell.value = result.result
        cell = sheet.cell(row=cell_row, column=7)
        cell = cell_style(cell)
        cell.value = result.researcher.get_full_name()
        cell = sheet.cell(row=cell_row, column=8)
        cell = cell_style(cell)
        cell.value = 'Да' if result.is_processed else 'Нет'
        cell_row += 1
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        with open(tmp.name, 'rb') as f:
            f.seek(0)
            xlsx_data = f.read()
    response = HttpResponse(xlsx_data, content_type=('application/vnd.openxml'
                            + 'formats-officedocument.spreadsheetml.sheet'))
    if marked_results:
        name = f'results_table_{int(time())}.xlsx'
    else:
        name = f'results_table_{datetime.now().strftime("%d.%m.%Y")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={name}'
    return response
